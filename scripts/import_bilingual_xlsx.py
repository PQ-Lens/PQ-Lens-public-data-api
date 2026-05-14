#!/usr/bin/env python3
"""Import bilingual.xlsx Maltese-English rows into the dataset API.

The script is intentionally idempotent: existing datasets and records are treated
as already imported, and the final verification step confirms that every row in
bilingual.xlsx is available through the API with the expected metadata.
"""

from __future__ import annotations

import argparse
import json
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Any
from urllib.error import HTTPError, URLError
from urllib.parse import quote, urlencode
from urllib.request import Request, urlopen

from openpyxl import load_workbook

DATASET_ID = "court_orders_bilingual"
SOURCE_LANGUAGE = "eng_Latn"
TARGET_LANGUAGE = "mlt_Latn"
LANGUAGE_PAIR = f"{SOURCE_LANGUAGE}-{TARGET_LANGUAGE}"
WORKSHEET_NAME = "Bilingual"
EXPECTED_HEADERS = ("English", "Maltese")
DEFAULT_BASE_URL = "http://localhost:5000"

DATASET_SUMMARY = (
    "A bilingual legal-text dataset derived from Court Notices published on the "
    "Government of Malta website. It contains parallel Maltese–English text pairs "
    "extracted from court notice PDFs and converted into structured text format."
)
DATASET_DESCRIPTION = (
    "This dataset contains Court Notices from the Government of Malta website "
    "(gov.mt). It contains Maltese–English pairs for each court notice. The "
    "original dataset format was PDF text; these extracts have been converted "
    "into structured text pairs in both Maltese and English. The dataset contains "
    "2,310 rows of data. The latest record date is 7 May 2026, and the earliest "
    "record date is 22 November 2022. The dataset is suitable for Maltese–English "
    "and English–Maltese machine translation tasks, legal NLP research tasks, and "
    "bilingual legal language analysis."
)


class ImportErrorWithDetails(Exception):
    """Raised when import or verification cannot continue safely."""


@dataclass(frozen=True)
class BilingualRow:
    row_number: int
    english_text: str
    maltese_text: str

    @property
    def record_id(self) -> str:
        return f"bilingual_row_{self.row_number:04d}"


def normalize_cell(value: Any) -> str:
    if value is None:
        return ""
    text = str(value)
    lines = [line.rstrip() for line in text.splitlines()]
    return "\n".join(lines).strip()


def load_rows(path: Path) -> list[BilingualRow]:
    if not path.exists():
        raise ImportErrorWithDetails(f"Excel file not found: {path}")

    workbook = load_workbook(path, read_only=True, data_only=True)
    if WORKSHEET_NAME not in workbook.sheetnames:
        raise ImportErrorWithDetails(
            f"Worksheet '{WORKSHEET_NAME}' not found in {path}. "
            f"Available sheets: {', '.join(workbook.sheetnames)}"
        )

    worksheet = workbook[WORKSHEET_NAME]
    headers = tuple(normalize_cell(worksheet.cell(row=1, column=idx).value) for idx in range(1, 3))
    if headers != EXPECTED_HEADERS:
        raise ImportErrorWithDetails(
            f"Expected headers {EXPECTED_HEADERS}, found {headers} in {path}"
        )

    rows: list[BilingualRow] = []
    for row_number in range(2, worksheet.max_row + 1):
        english_text = normalize_cell(worksheet.cell(row=row_number, column=1).value)
        maltese_text = normalize_cell(worksheet.cell(row=row_number, column=2).value)
        if not english_text and not maltese_text:
            continue
        if not english_text or not maltese_text:
            print(
                f"Skipping row {row_number}: both English and Maltese text are required for a bilingual pair",
                file=sys.stderr,
            )
            continue
        rows.append(
            BilingualRow(
                row_number=row_number,
                english_text=english_text,
                maltese_text=maltese_text,
            )
        )
    return rows


def request_json(
    method: str,
    base_url: str,
    path: str,
    *,
    payload: dict[str, Any] | None = None,
    query: dict[str, Any] | None = None,
) -> tuple[int, dict[str, Any]]:
    url = f"{base_url.rstrip('/')}{path}"
    if query:
        url = f"{url}?{urlencode(query)}"

    body = None if payload is None else json.dumps(payload).encode("utf-8")
    request = Request(url, data=body, method=method)
    request.add_header("Accept", "application/json")
    if payload is not None:
        request.add_header("Content-Type", "application/json")

    try:
        with urlopen(request, timeout=30) as response:
            response_body = response.read().decode("utf-8")
            return response.status, json.loads(response_body) if response_body else {}
    except HTTPError as exc:
        response_body = exc.read().decode("utf-8")
        try:
            parsed = json.loads(response_body) if response_body else {}
        except json.JSONDecodeError:
            parsed = {"error": response_body}
        return exc.code, parsed
    except URLError as exc:
        raise ImportErrorWithDetails(f"Could not reach API at {base_url}: {exc.reason}") from exc


def dataset_payload() -> dict[str, Any]:
    return {
        "id": DATASET_ID,
        "summary": DATASET_SUMMARY,
        "description": DATASET_DESCRIPTION,
        "language_mode": "bilingual_aligned",
        "synthetic_status": "non_synthetic",
        "synthetic_details": {},
        "provenance": {
            "source_url": "https://www.gov.mt/",
            "source_text": "Court Notices from the Government of Malta website exported from PDF text into bilingual.xlsx.",
            "source_type": "web",
        },
    }


def ensure_dataset(base_url: str) -> None:
    payload = dataset_payload()
    status, body = request_json("POST", base_url, "/datasets", payload=payload)
    if status == 201:
        print(f"Created dataset: {DATASET_ID}")
        return
    if status == 409:
        patch_payload = {key: value for key, value in payload.items() if key != "id"}
        patch_status, patch_body = request_json(
            "PATCH",
            base_url,
            f"/datasets/{quote(DATASET_ID)}",
            payload=patch_payload,
        )
        if patch_status == 200:
            print(f"Updated existing dataset metadata: {DATASET_ID}")
            return
        raise ImportErrorWithDetails(
            f"Dataset exists but metadata update failed with HTTP {patch_status}: {patch_body}"
        )
    raise ImportErrorWithDetails(f"Dataset creation failed with HTTP {status}: {body}")


def build_record_payload(row: BilingualRow, xlsx_path: Path) -> dict[str, Any]:
    return {
        "id": row.record_id,
        "text": row.english_text,
        "language": SOURCE_LANGUAGE,
        "language_pair": LANGUAGE_PAIR,
        "synthetic_status": "non_synthetic",
        "translation_metadata": {
            "source_language": SOURCE_LANGUAGE,
            "target_language": TARGET_LANGUAGE,
            "target_text": row.maltese_text,
            "alignment_type": "row_aligned",
            "source_column": "English",
            "target_column": "Maltese",
        },
        "provenance": {
            "source_text": f"{xlsx_path.name}, worksheet {WORKSHEET_NAME}, row {row.row_number}",
            "source_type": "file",
        },
        "attributes": {
            "excel_file": xlsx_path.name,
            "worksheet": WORKSHEET_NAME,
            "row_number": row.row_number,
            "original_source": "Government of Malta Court Notices",
            "original_source_url": "https://www.gov.mt/",
            "original_format": "pdf",
            "structured_format": "xlsx",
        },
    }


def insert_records(base_url: str, rows: list[BilingualRow], xlsx_path: Path) -> tuple[int, int]:
    created = 0
    existing = 0
    for row in rows:
        status, body = request_json(
            "POST",
            base_url,
            f"/datasets/{quote(DATASET_ID)}/records",
            payload=build_record_payload(row, xlsx_path),
        )
        if status == 201:
            created += 1
            continue
        if status == 409:
            existing += 1
            continue
        raise ImportErrorWithDetails(
            f"Failed to insert row {row.row_number} as {row.record_id}; HTTP {status}: {body}"
        )
    return created, existing


def fetch_all_records(base_url: str) -> list[dict[str, Any]]:
    cursor = 0
    records: list[dict[str, Any]] = []
    while True:
        status, body = request_json(
            "GET",
            base_url,
            f"/datasets/{quote(DATASET_ID)}/records",
            query={"batch_size": 10000, "cursor": cursor, "order": "natural"},
        )
        if status != 200:
            raise ImportErrorWithDetails(f"Failed to fetch records; HTTP {status}: {body}")
        records.extend(body.get("data", []))
        next_cursor = body.get("next_cursor")
        if next_cursor is None:
            return records
        cursor = int(next_cursor)


def verify_inserted(expected_rows: list[BilingualRow], actual_records: list[dict[str, Any]]) -> None:
    expected_by_id = {row.record_id: row for row in expected_rows}
    actual_by_id = {record.get("id"): record for record in actual_records}
    missing_ids = sorted(set(expected_by_id) - set(actual_by_id))
    if missing_ids:
        preview = ", ".join(missing_ids[:10])
        raise ImportErrorWithDetails(f"Missing {len(missing_ids)} expected records: {preview}")

    failures: list[str] = []
    for record_id, expected in expected_by_id.items():
        record = actual_by_id[record_id]
        translation_metadata = record.get("translation_metadata") or {}
        attributes = record.get("attributes") or {}
        checks = {
            "language": record.get("language") == SOURCE_LANGUAGE,
            "language_pair": record.get("language_pair") == LANGUAGE_PAIR,
            "synthetic_status": record.get("synthetic_status") == "non_synthetic",
            "text": record.get("text") == expected.english_text,
            "target_language": translation_metadata.get("target_language") == TARGET_LANGUAGE,
            "target_text": translation_metadata.get("target_text") == expected.maltese_text,
            "row_number": attributes.get("row_number") == expected.row_number,
        }
        bad_fields = [name for name, ok in checks.items() if not ok]
        if bad_fields:
            failures.append(f"{record_id}: {', '.join(bad_fields)}")

    if failures:
        preview = "; ".join(failures[:10])
        raise ImportErrorWithDetails(f"Verification failed for {len(failures)} records: {preview}")

    if len(actual_records) < len(expected_rows):
        raise ImportErrorWithDetails(
            f"API returned {len(actual_records)} records but {len(expected_rows)} rows were expected"
        )


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Import bilingual.xlsx Maltese-English court notice pairs into the dataset API."
    )
    parser.add_argument("--base-url", default=DEFAULT_BASE_URL, help=f"API base URL (default: {DEFAULT_BASE_URL})")
    parser.add_argument("--xlsx", default="bilingual.xlsx", type=Path, help="Path to bilingual.xlsx")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    rows = load_rows(args.xlsx)
    if not rows:
        raise ImportErrorWithDetails(f"No importable bilingual rows found in {args.xlsx}")

    ensure_dataset(args.base_url)
    created, existing = insert_records(args.base_url, rows, args.xlsx)
    records = fetch_all_records(args.base_url)
    verify_inserted(rows, records)

    print(
        "Import verified successfully: "
        f"{len(rows)} expected rows, {created} created, {existing} already existed, "
        f"{len(records)} records available in dataset '{DATASET_ID}'."
    )
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except ImportErrorWithDetails as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        raise SystemExit(1) from exc
