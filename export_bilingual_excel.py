#!/usr/bin/env python3
"""Export first/second page text from PDFs into an Excel sheet.

Column A: English (page 1)
Column B: Maltese (page 2)
"""

import argparse
import os
import sys


def iter_pdf_files(root, recursive=True, include_hidden=False):
    if recursive:
        for dirpath, dirnames, filenames in os.walk(root):
            if not include_hidden:
                dirnames[:] = [d for d in dirnames if not d.startswith(".")]
            for name in filenames:
                if not include_hidden and name.startswith("."):
                    continue
                if name.lower().endswith(".pdf"):
                    yield os.path.join(dirpath, name)
    else:
        for name in os.listdir(root):
            if not include_hidden and name.startswith("."):
                continue
            if name.lower().endswith(".pdf"):
                yield os.path.join(root, name)


def load_pypdf():
    try:
        from pypdf import PdfReader  # type: ignore
    except Exception as exc:
        print(
            "Missing dependency 'pypdf'. Install in your venv with:\n"
            "  . .venv/bin/activate && python -m pip install pypdf\n",
            file=sys.stderr,
        )
        raise SystemExit(2) from exc
    return PdfReader


def load_openpyxl():
    try:
        from openpyxl import Workbook  # type: ignore
        from openpyxl.styles import Alignment  # type: ignore
    except Exception as exc:
        print(
            "Missing dependency 'openpyxl'. Install in your venv with:\n"
            "  . .venv/bin/activate && python -m pip install openpyxl\n",
            file=sys.stderr,
        )
        raise SystemExit(2) from exc
    return Workbook, Alignment


def normalize_text(text):
    if text is None:
        return ""
    # Keep line breaks but trim trailing spaces on each line.
    lines = [line.rstrip() for line in text.splitlines()]
    return "\n".join(lines).strip()


def main():
    parser = argparse.ArgumentParser(
        description="Export English/Maltese page text from PDFs into Excel"
    )
    parser.add_argument(
        "--dir",
        default=os.getcwd(),
        help="Directory to scan (default: current directory)",
    )
    parser.add_argument(
        "--out",
        default="bilingual.xlsx",
        help="Output Excel filename (relative to --dir)",
    )
    parser.add_argument(
        "--include-filename",
        action="store_true",
        help="Add a leading column with the PDF filename",
    )
    parser.add_argument(
        "--no-recursive",
        action="store_true",
        help="Do not scan subdirectories",
    )
    parser.add_argument(
        "--include-hidden",
        action="store_true",
        help="Include hidden files/directories",
    )
    args = parser.parse_args()

    root = os.path.abspath(args.dir)
    out_path = args.out
    if not os.path.isabs(out_path):
        out_path = os.path.join(root, out_path)

    PdfReader = load_pypdf()
    Workbook, Alignment = load_openpyxl()

    pdf_files = sorted(
        iter_pdf_files(root, recursive=not args.no_recursive, include_hidden=args.include_hidden)
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Bilingual"

    headers = []
    if args.include_filename:
        headers.append("File")
    headers.extend(["English", "Maltese"])
    ws.append(headers)

    wrap = Alignment(wrap_text=True, vertical="top")

    for path in pdf_files:
        rel_path = os.path.relpath(path, root)
        try:
            reader = PdfReader(path)
            page1 = reader.pages[0].extract_text() if len(reader.pages) > 0 else ""
            page2 = reader.pages[1].extract_text() if len(reader.pages) > 1 else ""
        except Exception as exc:
            print(f"Error reading {rel_path}: {exc}", file=sys.stderr)
            page1 = ""
            page2 = ""

        row = []
        if args.include_filename:
            row.append(rel_path)
        row.append(normalize_text(page1))
        row.append(normalize_text(page2))
        ws.append(row)

    # Apply formatting
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = wrap

    # Set column widths to readable defaults
    start_col = 1
    if args.include_filename:
        ws.column_dimensions["A"].width = 48
        start_col = 2
    ws.column_dimensions[chr(64 + start_col)].width = 80
    ws.column_dimensions[chr(64 + start_col + 1)].width = 80

    ws.freeze_panes = "A2"

    wb.save(out_path)
    print(f"Wrote: {out_path}")
    print(f"Rows: {len(pdf_files)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
